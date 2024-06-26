# Jack Estes, Katelyn Hamilton, Hannah Larkin, Sebastian Mcfarland, Allen Schultz, Hayden Whalen
# IS303 Section 3
#
# Project 1 - Take poorly formatted input data from excel sheets and create a more organized excel sheet w/
# different workbooks & statistics
import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


# Create each sheet's structure within the workbook
def startPage(name, myBook):
    # Headers
    currSheet = myBook[name]
    currSheet.cell(1, 1).value = "Last Name"
    currSheet.column_dimensions["A"].width = len(currSheet.cell(1, 1).value) + 5
    currSheet.cell(1, 1).font = Font(bold=True)
    currSheet.cell(1, 2).value = "First Name"
    currSheet.column_dimensions["B"].width = len(currSheet.cell(1, 2).value) + 5
    currSheet.cell(1, 2).font = Font(bold=True)
    currSheet.cell(1, 3).value = "Student ID"
    currSheet.column_dimensions["C"].width = len(currSheet.cell(1, 3).value) + 5
    currSheet.cell(1, 3).font = Font(bold=True)
    currSheet.cell(1, 4).value = "Grade"
    currSheet.column_dimensions["D"].width = len(currSheet.cell(1, 4).value) + 5
    currSheet.cell(1, 4).font = Font(bold=True)
    currSheet.cell(1, 6).value = "Summary Statistics"
    currSheet.column_dimensions["F"].width = len(currSheet.cell(1, 6).value) + 5
    currSheet.cell(1, 6).font = Font(bold=True)
    currSheet.cell(1, 7).value = "Value"
    currSheet.column_dimensions["G"].width = len(currSheet.cell(1, 7).value) + 5
    currSheet.cell(1, 7).font = Font(bold=True)

    # Filter data
    currSheet.auto_filter.ref = "A:D"


# Create a sheet for each unique class in the poorly organized data in an excel file
def createSheets(poorData, classesList, goodData):
    for row in poorData.iter_rows(min_row=2):
        # Add the class name to the classesList if it isn't already in there
        if not classesList.__contains__(row[0].value):
            classesList.append(row[0].value)
            goodData.create_sheet(row[0].value)
            startPage(row[0].value, goodData)


# Organizes the data into each respective class
def parseData(poorData, goodData):
    # In poor data:
    # row[0] is class name
    # row[1] is student info (firstName_lastName_ID)
    # row[2] is grade value
    for row in poorData:
        studentInfoList = row[1].value.split("_")
        if len(studentInfoList) > 1:
            firstName = studentInfoList[0]
            lastName = studentInfoList[1]
            studentID = studentInfoList[2]
            studentGrade = row[2].value
            goodData[row[0].value].append([lastName, firstName, studentID, studentGrade])


# Sets class statistics within each sheet of the workbook
def setGrades(currSheet):
    # Non Headers
    currSheet.cell(2, 6).value = "Highest Grade"
    currSheet.cell(3, 6).value = "Lowest Grade"
    currSheet.cell(4, 6).value = "Mean Grade"
    currSheet.cell(5, 6).value = "Median Grade"
    currSheet.cell(6, 6).value = "Number of Students"
    currSheet["G2"] = '=MAX(OFFSET($D$2,0,0,COUNT($D:$D)-1,1))'
    currSheet["G3"] = '=MIN(OFFSET($D$2,0,0,COUNT($D:$D)-1,1))'
    currSheet["G4"] = '=IF(G6=0,0,AVERAGE(OFFSET($D$2,0,0,COUNT($D:$D)-1,1)))'
    currSheet["G5"] = '=IF(G6=0,0,MEDIAN(OFFSET($D$2,0,0,COUNT($D:$D)-1,1)))'
    currSheet["G6"] = '=COUNT(OFFSET($D$2,0,0,COUNT($D:$D)-1,1))'


# Create objects for poorly organized spreadsheets w/ our data
poorDataXlsx1 = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")
poorDataXlsx2 = openpyxl.load_workbook("Poorly_Organized_Data_2.xlsx")

# Create objects for well organized spreadsheet data
goodDataXlsx1 = Workbook()
goodDataXlsx2 = Workbook()

# Create sheets for each class in poorDataXlsx1 & poorDataXlsx2
createSheets(poorDataXlsx1.active, [], goodDataXlsx1)
createSheets(poorDataXlsx2.active, [], goodDataXlsx2)

# Remove default sheets if they exist
goodDataXlsx1.__delitem__("Sheet")
goodDataXlsx2.__delitem__("Sheet")

# Parse data into each sheet
parseData(poorDataXlsx1.active, goodDataXlsx1)
parseData(poorDataXlsx2.active, goodDataXlsx2)

# Set the grade statistics values for each sheet
for sheet in goodDataXlsx1.sheetnames:
    setGrades(goodDataXlsx1[sheet])
for sheet in goodDataXlsx2.sheetnames:
    setGrades(goodDataXlsx2[sheet])

# Save the well organized data
outputFile1 = "Formatted_Grades_1.xlsx"
outputFile2 = "Formatted_Grades_2.xlsx"
goodDataXlsx1.save(filename=outputFile1)
goodDataXlsx2.save(filename=outputFile2)

# Close all excel files used to avoid resource leak
goodDataXlsx1.close()
goodDataXlsx2.close()
poorDataXlsx1.close()
poorDataXlsx2.close()
os.system(f"open -a '/Applications/Microsoft/Microsoft Excel.app' {outputFile1}")
os.system(f"open -a '/Applications/Numbers.app' {outputFile2}")
