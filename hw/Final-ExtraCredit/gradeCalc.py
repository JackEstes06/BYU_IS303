# Jack Estes
# 4/11/24
# IS303 Section 3
# Extra Credit - recreate an excel spreadsheet like the GradeCalc.xlsx in the folder
import os

import openpyxl
from openpyxl import styles
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import NumberFormatDescriptor


# Set column widths for sheet
def sizeColumns(sheet):
    sheet.column_dimensions["A"].width = 21.83
    sheet.column_dimensions["A"].font = openpyxl.styles.Font(bold=True)
    sheet.column_dimensions["B"].width = 8
    sheet.column_dimensions["C"].width = 8
    sheet.column_dimensions["D"].width = 7.83


# Set the colors for a given range
def setColors(sheet, startRow, endRow, startCol, endCol):
    for row in range(startRow, endRow + 1):
        for col in range(startCol, endCol + 1):
            sheet.cell(row, col).fill = openpyxl.styles.fills.PatternFill(start_color="d9d9d9", fill_type="solid")


# Given a specific row, startCol, and endCol put a border around these cells
def setBorders(sheet, startRow, endRow, startCol, endCol):
    cellBorderStyle = openpyxl.styles.borders.Side(border_style="thin")

    for row in range(startRow, endRow + 1):
        topBorder = False
        bottomBorder = False
        if row == startRow:
            topBorder = True
        if row == endRow:
            bottomBorder = True
        for col in range(startCol, endCol + 1):
            leftBorder = False
            rightBorder = False
            if col == startCol:
                leftBorder = True
            if col == endCol:
                rightBorder = True
            sheet.cell(row, col).border = openpyxl.styles.borders.Border(
                left=cellBorderStyle if leftBorder else None,
                right=cellBorderStyle if rightBorder else None,
                top=cellBorderStyle if topBorder else None,
                bottom=cellBorderStyle if bottomBorder else None,
            )
    # sheet.merge_cells("A1:D1")
    # sheet.unmerge_cells("A1:D1")


def addWords(sheet, cellRow, cellCol, words, isBold):
    sheet.cell(cellRow, cellCol).value = words
    sheet.cell(cellRow, cellCol).font = Font(bold=isBold)


gradeCalc = openpyxl.Workbook()
# There is only 1 sheet so we just get the first worksheet as the currSheet
currSheet = gradeCalc[gradeCalc.sheetnames[0]]

# Set the styling for the sheet
sizeColumns(currSheet)
setBorders(currSheet, 1, 1, 1, 4)
setBorders(currSheet, 3, 5, 1, 4)
setBorders(currSheet, 7, 8, 1, 4)
setBorders(currSheet, 10, 11, 1, 4)
setBorders(currSheet, 13, 13, 2, 4)
setBorders(currSheet, 15, 15, 1, 4)
setBorders(currSheet, 17, 17, 1, 4)
currSheet.merge_cells("B13:C13")
currSheet.merge_cells("A17:C17")
setColors(currSheet, 17, 17, 1, 1)
setColors(currSheet, 13, 13, 2, 2)
setColors(currSheet, 7, 7, 3, 4)

# Set the words to the sheet & style them
addWords(currSheet, 1, 1, "Assignments (Drop lowest)", True)
addWords(currSheet, 3, 1, "Project 1", True)
addWords(currSheet, 4, 1, "Project 2", True)
addWords(currSheet, 5, 1, "Project 3", True)
addWords(currSheet, 7, 1, "Midterm", True)
addWords(currSheet, 8, 1, "Midterm Extra Credit", True)
addWords(currSheet, 10, 1, "Final", True)
addWords(currSheet, 11, 1, "Final Exam Extra Credit", True)
addWords(currSheet, 13, 2, "Calculated Grade", True)
addWords(currSheet, 15, 1, "SONA", True)
addWords(currSheet, 17, 1, "Final Grade without Curve", True)
currSheet.cell(13, 2).alignment = Alignment(horizontal="right")
currSheet.cell(17, 1).alignment = Alignment(horizontal="right")

# Set the weigh values for grading calculations
currSheet["C1"] = 0.15
currSheet["C3"] = 0.1
currSheet["C4"] = 0.1
currSheet["C5"] = 0.15
currSheet["C8"] = 0.2
currSheet["C11"] = 0.3

# Gather inputs of each grade from user
currSheet["B1"] = float(input("What is your % out of 100 for Assignments? "))
currSheet["B3"] = float(input("What is your % out of 100 for Project 1? "))
currSheet["B4"] = float(input("What is your % out of 100 for Project 2? "))
currSheet["B5"] = float(input("What is your % out of 100 for Project 3? "))
currSheet["B7"] = float(input("What is your % out of 100 for Midterm? "))
currSheet["B8"] = float(input("How many extra credit points did you earn for the Midterm? "))
currSheet["B10"] = float(input("What is your % out of 100 for Final? "))
currSheet["B11"] = float(input("How many extra credit points did you earn for the Final? "))
currSheet["B15"] = float(input("How much SONA credit did you earn? "))

# Calculate the values based on inputs
currSheet["D1"] = "=B1*C1"
currSheet["D3"] = "=B3*C3"
currSheet["D4"] = "=B4*C4"
currSheet["D5"] = "=B5*C5"
currSheet["D8"] = "=(B7+B8)*C8"
currSheet["D11"] = "=(B10+B11)*C11"
currSheet["D13"] = "=SUM(D1:D11)"
currSheet["D17"] = "=D13+B15"

# Format numbers
currSheet.cell(1, 4).number_format = '0.0'
currSheet.cell(3, 4).number_format = '0.0'
currSheet.cell(4, 4).number_format = '0.0'
currSheet.cell(5, 4).number_format = '0.0'
currSheet.cell(8, 4).number_format = '0.0'
currSheet.cell(11, 4).number_format = '0.0'
currSheet.cell(13, 4).number_format = '0.0'
currSheet.cell(17,4).font = Font(bold=True)
currSheet.cell(17, 4).number_format = '0.0'

# Save the new workbook
outputFile = "GradeCalcMimic.xlsx"
gradeCalc.save(filename=outputFile)

# Close all excel files to avoid resource leaks
gradeCalc.close()
os.system(f"open -a '/Applications/Microsoft/Microsoft Excel.app' {outputFile}")
